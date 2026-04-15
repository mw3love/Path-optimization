#!/usr/bin/env python3
"""
xlsx_to_locations.py
xlsx 또는 구글시트(CSV export) → locations.json 변환 스크립트

사용 예시:
  python tools/xlsx_to_locations.py --xlsx "측정표(배포용)_전북.xlsx" --out locations.json
  python tools/xlsx_to_locations.py --gid 0 --sheet-id 1cStO5DrZoFDOmZibuycWXapNuDLR1IWNwrnQh2RFFqo --out locations.json

헤더 스키마 (Row 3):
  NA | Point_Name | Address_1 | Address_2 | Address_3 | Address_4 | Latitude | Longitude | ...

출력 JSON 스키마:
  {"id": str, "seq": int, "name": str, "address": str, "sigungu": str, "lat": float, "lng": float}
"""

import argparse
import csv
import io
import json
import sys
import urllib.request
from typing import Optional


REQUIRED_HEADERS = {"NA", "Point_Name", "Address_4", "Latitude", "Longitude"}


def parse_args():
    parser = argparse.ArgumentParser(
        description="xlsx/구글시트 → locations.json 변환"
    )
    source = parser.add_mutually_exclusive_group(required=True)
    source.add_argument("--xlsx", metavar="PATH", help="로컬 xlsx 파일 경로")
    source.add_argument(
        "--sheet-id",
        metavar="SHEET_ID",
        help="구글시트 ID (--gid 와 함께 사용)",
    )
    parser.add_argument(
        "--gid",
        metavar="GID",
        default="0",
        help="구글시트 탭 gid (기본값: 0)",
    )
    parser.add_argument(
        "--sheet",
        metavar="SHEET_NAME",
        default=None,
        help="xlsx 시트 이름 (생략 시 첫 번째 시트)",
    )
    parser.add_argument(
        "--out",
        metavar="PATH",
        default="locations.json",
        help="출력 JSON 파일 경로 (기본값: locations.json)",
    )
    parser.add_argument(
        "--strict",
        action="store_true",
        help="엄격 모드: 오류 행 발생 시 즉시 종료",
    )
    return parser.parse_args()


def find_header_row(rows: list[list]) -> Optional[int]:
    """'NA' 와 'Point_Name' 이 있는 헤더 행 인덱스(0-based) 반환"""
    for i, row in enumerate(rows):
        str_cells = [str(c).strip() if c is not None else "" for c in row]
        if "NA" in str_cells and "Point_Name" in str_cells:
            return i
    return None


def build_col_map(header_row: list) -> dict[str, int]:
    """헤더 행 → {컬럼명: 인덱스} 매핑"""
    return {
        str(cell).strip(): idx
        for idx, cell in enumerate(header_row)
        if cell is not None and str(cell).strip()
    }


def validate_col_map(col_map: dict[str, int]) -> list[str]:
    """필수 컬럼 누락 목록 반환"""
    return [h for h in REQUIRED_HEADERS if h not in col_map]


def parse_row(
    row: list,
    col_map: dict[str, int],
    row_num: int,
    strict: bool,
) -> Optional[dict]:
    """단일 데이터 행 → dict. 오류 시 strict=True면 예외, 아니면 None 반환"""

    def get(col: str):
        idx = col_map.get(col)
        if idx is None or idx >= len(row):
            return None
        v = row[idx]
        return str(v).strip() if v is not None else None

    na = get("NA")
    name = get("Point_Name")
    address = get("Address_4")
    sigungu = get("Address_2") or ""
    lat_str = get("Latitude")
    lng_str = get("Longitude")

    # NA(seq) 필수
    if not na:
        msg = f"row {row_num}: NA(seq) 값이 없음 — 스킵"
        if strict:
            raise ValueError(msg)
        print(f"[WARN] {msg}", file=sys.stderr)
        return None

    # 이름 필수
    if not name:
        msg = f"row {row_num}: Point_Name 값이 없음 — 스킵 (NA={na})"
        if strict:
            raise ValueError(msg)
        print(f"[WARN] {msg}", file=sys.stderr)
        return None

    # 좌표 필수 + 숫자 검증
    try:
        lat = float(lat_str) if lat_str else None
        lng = float(lng_str) if lng_str else None
    except (ValueError, TypeError):
        lat = lng = None

    if lat is None or lng is None:
        msg = f"row {row_num}: 좌표 누락/형식 오류 (NA={na}, lat={lat_str!r}, lng={lng_str!r}) — 스킵"
        if strict:
            raise ValueError(msg)
        print(f"[WARN] {msg}", file=sys.stderr)
        return None

    # NA 값을 seq(정수)와 id(문자열) 로 분리
    try:
        seq = int(float(na))
    except (ValueError, TypeError):
        seq = 0

    loc_id = str(int(float(na))) if seq else str(na)

    return {
        "id": loc_id,
        "seq": seq,
        "name": name,
        "address": address or "",
        "sigungu": sigungu,
        "lat": lat,
        "lng": lng,
    }


def rows_from_xlsx(path: str, sheet_name: Optional[str]) -> list[list]:
    try:
        import openpyxl
    except ImportError:
        print("[ERROR] openpyxl 패키지가 없습니다: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(path, data_only=True)
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            print(
                f"[ERROR] 시트 '{sheet_name}' 을 찾을 수 없습니다. "
                f"사용 가능한 시트: {wb.sheetnames}",
                file=sys.stderr,
            )
            sys.exit(1)
        ws = wb[sheet_name]
    else:
        ws = wb.active

    return [list(row) for row in ws.iter_rows(values_only=True)]


def rows_from_google_sheet(sheet_id: str, gid: str) -> list[list]:
    url = (
        f"https://docs.google.com/spreadsheets/d/{sheet_id}"
        f"/export?format=csv&gid={gid}"
    )
    print(f"[INFO] 구글시트 다운로드: {url}", file=sys.stderr)
    try:
        with urllib.request.urlopen(url, timeout=30) as resp:
            content = resp.read().decode("utf-8")
    except Exception as e:
        print(f"[ERROR] 구글시트 다운로드 실패: {e}", file=sys.stderr)
        sys.exit(1)

    reader = csv.reader(io.StringIO(content))
    return [row for row in reader]


def convert(rows: list[list], strict: bool) -> list[dict]:
    header_idx = find_header_row(rows)
    if header_idx is None:
        print(
            "[ERROR] 헤더 행(NA + Point_Name 포함)을 찾을 수 없습니다.",
            file=sys.stderr,
        )
        sys.exit(1)

    col_map = build_col_map(rows[header_idx])
    missing = validate_col_map(col_map)
    if missing:
        print(
            f"[ERROR] 필수 컬럼 누락: {missing}",
            file=sys.stderr,
        )
        sys.exit(1)

    locations = []
    skipped = 0
    for i, row in enumerate(rows[header_idx + 1 :], start=header_idx + 2):
        # 완전히 빈 행 스킵
        if all(c is None or str(c).strip() == "" for c in row):
            continue
        result = parse_row(row, col_map, i, strict)
        if result is None:
            skipped += 1
        else:
            locations.append(result)

    if skipped:
        print(f"[INFO] {skipped}개 행 스킵됨", file=sys.stderr)

    return locations


def main():
    args = parse_args()

    if args.xlsx:
        rows = rows_from_xlsx(args.xlsx, args.sheet)
    else:
        rows = rows_from_google_sheet(args.sheet_id, args.gid)

    locations = convert(rows, args.strict)

    if not locations:
        print("[ERROR] 변환된 지점이 0개입니다.", file=sys.stderr)
        sys.exit(1)

    with open(args.out, "w", encoding="utf-8") as f:
        json.dump(locations, f, ensure_ascii=False, indent=2)

    print(
        f"[OK] {len(locations)}개 지점 → {args.out}",
        file=sys.stderr,
    )


if __name__ == "__main__":
    main()
